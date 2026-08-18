from __future__ import annotations
import argparse
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import Request, urlopen
from openpyxl import load_workbook

from lead_studio.card_files import render_brief
from lead_studio.website_repair import repair_missing_website_data as repair_websites_from_sources


ROOT = Path(__file__).parent.resolve()
CONFIG_PATH = ROOT.parent / "config.json"
try:
    CONFIG = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
except Exception:
    CONFIG = {"parser": {"headers": {}, "social_domains": [], "max_photos": 12}}

HEADERS = CONFIG["parser"].get("headers", {})
DEFAULT_SOCIAL_DOMAINS = (
    "yclients.com", "dikidi.net", "dikidi.ru", "prodoctorov.ru", "zoon.ru",
    "vk.com", "vk.ru", "max.ru", "t.me", "wa.me", "whatsapp.com",
    "instagram.com", "facebook.com", "viber.com", "youtube.com", "ok.ru",
    "taplink.cc", "aqulas.me", "nethouse.ru",
)
SOCIAL_DOMAINS = tuple(dict.fromkeys((*DEFAULT_SOCIAL_DOMAINS, *CONFIG["parser"].get("social_domains", []))))
IMAGE_SIZE = "L_height"

def text(value: Any) -> str:
    return "" if value is None else str(value).strip()

def read_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = [text(value) for value in next(rows, ())]
    result = []
    for row in rows:
        values = [text(value) for value in row]
        if any(values):
            result.append(dict(zip(header, values)))
    return result

def first_value(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key, "")
        if value:
            return value
    return ""


def normalize_coords(coords: list, row_value: str) -> list[float]:
    if coords:
        return coords
    parts = re.split(r"[,;\s]+", row_value.strip())
    result = []
    for part in parts:
        try:
            result.append(float(part.replace(",", ".")))
        except ValueError:
            continue
    return result


def extract_state(html: str) -> dict[str, Any]:
    for match in re.finditer(r"<script[^>]*>(.*?)</script>", html, flags=re.S):
        value = match.group(1).strip()
        if value.startswith("{") and '"stack"' in value and '"config"' in value:
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                continue
    raise ValueError("Yandex state JSON not found")


def http_get_html(url: str) -> str:
    request = Request(url, headers=HEADERS)
    with urlopen(request, timeout=30) as response:
        return response.read().decode("utf-8", errors="replace")


def fetch_org_item(url: str) -> dict[str, Any]:
    html = http_get_html(url)
    state = extract_state(html)
    stack = state.get("stack") or []
    for entry in stack:
        items = ((entry.get("results") or {}).get("items")) or []
        for item in items:
            if item.get("type") == "business":
                return item
    raise ValueError("Business item not found")


def image_urls(item: dict[str, Any], limit: int) -> list[dict[str, str]]:
    photos = ((item.get("photos") or {}).get("items")) or []
    result = []
    for photo in photos[:limit]:
        template = text(photo.get("urlTemplate"))
        if not template:
            continue
        result.append(
            {
                "url": template.replace("%s", IMAGE_SIZE),
                "template": template,
                "alt": text(photo.get("alt")),
            }
        )
    return result


def feature_names(item: dict[str, Any]) -> list[str]:
    result = []
    for feature in item.get("features") or []:
        name = text(feature.get("name"))
        value = feature.get("value")
        if not name:
            continue
        if value is True:
            result.append(name)
        elif isinstance(value, str):
            result.append(f"{name}: {value}")
        elif isinstance(value, list):
            nested = []
            for v in value:
                if isinstance(v, dict):
                    nested.append(text(v.get("name")))
                else:
                    nested.append(str(v).strip())
            nested = [v for v in nested if v]
            if nested:
                result.append(f"{name}: {', '.join(nested)}")
    return result


def links_from_item(item: dict[str, Any], row: dict[str, str]) -> tuple[list[str], list[str]]:
    websites = []
    socials = []

    def is_host_in(host: str, domains: tuple[str, ...]) -> bool:
        return any(host == domain or host.endswith(f".{domain}") for domain in domains)

    def unwrap_vk_away(href: str) -> str:
        parsed = urlparse(href)
        host = (parsed.hostname or "").lower()
        if host in {"vk.ru", "vk.com"} and parsed.path.startswith("/away"):
            return unquote(parse_qs(parsed.query).get("to", [""])[0]).strip()
        return href

    def href_host(href: str) -> str:
        parsed = urlparse(href)
        if not parsed.scheme:
            parsed = urlparse(f"https:{href}" if href.startswith("//") else f"https://{href}")
        if parsed.scheme not in {"http", "https"}:
            return ""
        return (parsed.hostname or "").lower()

    def social_key(href: str) -> str:
        parsed = urlparse(href)
        if not parsed.scheme:
            parsed = urlparse(f"https:{href}" if href.startswith("//") else f"https://{href}")
        host = (parsed.hostname or "").lower()
        path = parsed.path.strip("/")
        if is_host_in(host, ("t.me", "telegram.me", "telegram.org")):
            account = path.lstrip("+")
            if re.fullmatch(r"\d{7,15}", account):
                return f"telegram:{account}"
        if is_host_in(host, ("wa.me", "whatsapp.com")):
            phone = re.sub(r"\D", "", parse_qs(parsed.query).get("phone", [path])[0])
            if phone:
                return f"whatsapp:{phone}"
        if is_host_in(host, ("vk.ru", "vk.com", "vkontakte.ru")):
            return f"vk:{path.rstrip('/').lower()}"
        return href

    def dedupe_socials(values: list[str]) -> list[str]:
        seen = set()
        result = []
        for href in values:
            key = social_key(href)
            if key in seen:
                continue
            seen.add(key)
            result.append(href)
        return result

    def route_link(href: str) -> None:
        href = unwrap_vk_away(text(href))
        host = href_host(href)
        if not host or is_host_in(host, ("yandex.ru", "ya.ru")):
            return
        if is_host_in(host, SOCIAL_DOMAINS):
            socials.append(href)
        else:
            websites.append(href)

    excel_site = first_value(row, "Сайт")
    route_link(excel_site)

    for link in list(item.get("businessLinks") or []) + [{"href": url} for url in item.get("urls") or []]:
        route_link(text(link.get("href") or link.get("url")))
    for link in item.get("socialLinks") or []:
        href = text(link.get("href"))
        unwrapped = unwrap_vk_away(href)
        if unwrapped != href:
            route_link(unwrapped)
        elif href:
            socials.append(href)
    excel_socials = first_value(row, "Соцсети", "whatsapp", "telegram", "vkontakte")
    if excel_socials:
        socials.extend([part.strip() for part in excel_socials.split(",") if part.strip()])
    return sorted(set(websites)), dedupe_socials(socials)


def repair_missing_website_data(repo: Any) -> int:
    return repair_websites_from_sources(repo, links_from_item)


def row_reviews(reviews: list[dict[str, str]], org_id: str, limit: int) -> list[dict[str, str]]:
    def review_date_key(item: dict[str, str]) -> tuple[int, int, int]:
        parts = first_value(item, "Дата обновления").split(".")
        if len(parts) == 3:
            try:
                day, month, year = (int(p) for p in parts)
                return (year, month, day)
            except ValueError:
                pass
        return (-1, -1, -1)

    matched = [review for review in reviews if first_value(review, "ID бизнеса") == org_id]
    matched.sort(key=review_date_key, reverse=True)
    return [
        {
            "rating": first_value(review, "Оценка"),
            "text": first_value(review, "Текст"),
            "author": first_value(review, "Автор"),
            "date": first_value(review, "Дата обновления"),
            "source": first_value(review, "Ссылка Яндекс"),
        }
        for review in matched[:limit]
    ]


def enrich_row(
    row: dict[str, str],
    reviews: list[dict[str, str]],
    *,
    max_reviews: int,
    max_photos: int,
    no_network: bool,
) -> dict[str, Any]:
    org_id = first_value(row, "ID организации", "ID")
    card_url = first_value(row, "Ссылка на карточку") or f"https://yandex.ru/maps/org/{org_id}"
    item: dict[str, Any] = {}
    error = ""
    if not no_network:
        try:
            item = fetch_org_item(card_url)
        except Exception as exc:  # noqa: BLE001 - keep row-level error in output
            error = str(exc)

    websites, socials = links_from_item(item, row)
    rating = item.get("ratingData") or {}
    phones = item.get("phones") or []
    categories = item.get("categories") or []
    coords = item.get("coordinates") or []

    phones_from_row = [
        {"number": part.strip(), "info": ""}
        for part in first_value(row, "Контакты", "Телефон").split(",")
        if part.strip()
    ]
    return {
        "id": org_id,
        "name": text(item.get("title")) or first_value(row, "Название"),
        "category": ", ".join(text(c.get("name")) for c in categories if c.get("name"))
        or first_value(row, "Основная категория", "Подрубрика", "Рубрика"),
        "address": text(item.get("address")) or first_value(row, "Адрес"),
        "city": first_value(row, "Город"),
        "region": first_value(row, "Регион"),
        "coordinates": normalize_coords(coords, first_value(row, "Координаты", "Широта")),
        "rating": rating.get("ratingValue") or first_value(row, "Рейтинг"),
        "rating_count": rating.get("ratingCount") or first_value(row, "Оценок", "Кол-во оценок"),
        "review_count": rating.get("reviewCount") or first_value(row, "Отзывов", "Кол-во отзывов"),
        "phones": [
            {"number": text(phone.get("number")), "info": text(phone.get("info"))}
            for phone in phones
        ] or phones_from_row,
        "websites": websites,
        "socials": socials,
        "hours": text(item.get("workingTimeText")) or first_value(row, "Время работы"),
        "features": feature_names(item),
        "photos": image_urls(item, max_photos),
        "reviews": row_reviews(reviews, org_id, max_reviews),
        "yandex_url": card_url,
        "has_site": bool(websites),
        "fetch_error": error,
        "source_row": row,
    }


def write_outputs(leads: list[dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    briefs_dir = output_dir / "briefs"
    briefs_dir.mkdir(exist_ok=True)
    (output_dir / "leads.json").write_text(
        json.dumps(leads, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    for lead in leads:
        safe_id = re.sub(r"[^0-9A-Za-z_-]+", "_", lead["id"] or "unknown").strip("_") or "unknown"
        filename = f"{safe_id}.md"
        (briefs_dir / filename).write_text(render_brief(lead), encoding="utf-8-sig")


def main() -> int:
    parser = argparse.ArgumentParser(description="Build landing briefs from Yandex Maps Excel exports.")
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, default=Path("yamap_landing_output"))
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--max-reviews", type=int, default=8)
    parser.add_argument("--max-photos", type=int, default=12)
    parser.add_argument("--no-network", action="store_true")
    args = parser.parse_args()

    organizations = read_sheet_rows(args.input, "Организации")
    if not organizations:
        organizations = read_sheet_rows(args.input, "Sheet1")
    reviews = read_sheet_rows(args.input, "Отзывы")
    leads = [
        enrich_row(
            row,
            reviews,
            max_reviews=args.max_reviews,
            max_photos=args.max_photos,
            no_network=args.no_network,
        )
        for row in organizations[: args.limit]
    ]
    write_outputs(leads, args.output)
    print(f"organizations: {len(organizations)}")
    print(f"leads_written: {len(leads)}")
    print(f"output: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
